"""
test_pipeline.py
----------------
Testes unitários mockados para o pipeline Argentina Updater.
Não requerem arquivos reais, rede, SQL Server ou Excel instalado.

Executar:
    python -m pytest test_pipeline.py -v
    # ou
    python test_pipeline.py
"""
from __future__ import annotations

import io
import sys
import textwrap
import types
import unittest
from datetime import date, datetime
from pathlib import Path
from unittest.mock import MagicMock, Mock, patch, call

import pandas as pd


# ---------------------------------------------------------------------------
# Helpers para mockar módulos externos que podem não estar instalados
# ---------------------------------------------------------------------------

def _mock_module(name: str) -> types.ModuleType:
    mod = types.ModuleType(name)
    sys.modules[name] = mod
    return mod


# Garante que módulos externos não precisam estar instalados para os testes
if "pyodbc" not in sys.modules:
    _mock_module("pyodbc")
if "win32com" not in sys.modules:
    win32mod = _mock_module("win32com")
    _mock_module("win32com.client")
if "pythoncom" not in sys.modules:
    _mock_module("pythoncom")

# Selenium — mockado para que TestDownloadFile funcione sem Chrome instalado
if "selenium" not in sys.modules:
    sel = _mock_module("selenium")
    sel_wd = _mock_module("selenium.webdriver")
    sel_wd.Chrome = MagicMock
    sel_chrome_opts = _mock_module("selenium.webdriver.chrome.options")
    sel_chrome_opts.Options = MagicMock
    sel_chrome_svc = _mock_module("selenium.webdriver.chrome.service")
    sel_chrome_svc.Service = MagicMock
if "webdriver_manager" not in sys.modules:
    wdm = _mock_module("webdriver_manager")
    wdm_chrome = _mock_module("webdriver_manager.chrome")
    wdm_chrome.ChromeDriverManager = MagicMock


# ---------------------------------------------------------------------------
# Imports dos módulos do projeto
# (adicionamos src/ ao path se necessário)
# ---------------------------------------------------------------------------

sys.path.insert(0, str(Path(__file__).resolve().parent))

from database import (
    _cortar_apos_duas_linhas_vazias,
    ler_arquivo_novo,
    merge_com_banco,
    salvar_local,
    salvar_onedrive,
    salvar_sql_server,
)
from downloader import _build_output_name, _extract_server_filename, download_file
from latest_file import get_latest_file


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

def _make_df(rows: list[dict]) -> pd.DataFrame:
    return pd.DataFrame(rows)


def _sample_db() -> pd.DataFrame:
    return _make_df([
        {"Date": "2025-12-01", "Destination": "CHINA", "Origin": "ARGENTINA",
         "Cargo": "CORN", "Tons": 50000, "Month": 12, "Year": 2025},
        {"Date": "2025-12-15", "Destination": "INDIA", "Origin": "ARGENTINA",
         "Cargo": "SOYA", "Tons": 30000, "Month": 12, "Year": 2025},
        {"Date": "2026-01-10", "Destination": "CHINA", "Origin": "ARGENTINA",
         "Cargo": "CORN", "Tons": 45000, "Month": 1, "Year": 2026},
    ])


def _sample_novo() -> pd.DataFrame:
    return _make_df([
        {"Date": "2026-01-10", "Destination": "CHINA", "Origin": "ARGENTINA",
         "Cargo": "CORN", "Tons": 46000, "Month": 1, "Year": 2026},
        {"Date": "2026-01-20", "Destination": "INDIA", "Origin": "ARGENTINA",
         "Cargo": "CORN", "Tons": 20000, "Month": 1, "Year": 2026},
        {"Date": "2026-02-05", "Destination": "SPAIN", "Origin": "ARGENTINA",
         "Cargo": "CORN", "Tons": 15000, "Month": 2, "Year": 2026},
    ])


# ===========================================================================
# TESTES — database.py
# ===========================================================================

class TestCortarAposDuasLinhasVazias(unittest.TestCase):

    def test_sem_linhas_vazias_retorna_tudo(self):
        df = _make_df([{"A": 1}, {"A": 2}, {"A": 3}])
        result = _cortar_apos_duas_linhas_vazias(df)
        self.assertEqual(len(result), 3)

    def test_corta_apos_duas_linhas_vazias(self):
        df = _make_df([
            {"A": 1},
            {"A": 2},
            {"A": None},   # linha vazia 1
            {"A": None},   # linha vazia 2  → deve cortar aqui
            {"A": 999},    # rodapé — deve ser removido
        ])
        result = _cortar_apos_duas_linhas_vazias(df)
        self.assertEqual(len(result), 2)
        self.assertNotIn(999, result["A"].values)

    def test_uma_linha_vazia_nao_corta(self):
        df = _make_df([{"A": 1}, {"A": None}, {"A": 3}])
        result = _cortar_apos_duas_linhas_vazias(df)
        self.assertEqual(len(result), 3)


class TestLerArquivoNovo(unittest.TestCase):

    def test_ler_arquivo_novo(self):
        """Verifica que Date é convertido e Month/Year derivados corretamente."""
        df_raw = _make_df([
            {"Date": "2026-03-10", "Destination": "CHINA", "Origin": "ARGENTINA",
             "Cargo": "CORN", "Tons": 1000},
            {"Date": "2026-03-11", "Destination": "INDIA", "Origin": "ARGENTINA",
             "Cargo": "CORN", "Tons": 2000},
        ])

        with patch("database.pd.read_excel", return_value=df_raw), \
             patch("database._cortar_apos_duas_linhas_vazias", side_effect=lambda x: x):
            result = ler_arquivo_novo(Path("fake.xlsx"))

        self.assertEqual(result["Month"].iloc[0], 3)
        self.assertEqual(result["Year"].iloc[0], 2026)
        self.assertIsInstance(result["Date"].iloc[0], pd.Timestamp)


class TestMergeComBanco(unittest.TestCase):

    def test_remove_periodos_sobrepostos(self):
        db = _sample_db()
        novo = _sample_novo()  # tem Jan e Feb 2026
        result = merge_com_banco(novo, db)

        # Janeiro 2026 original (45000) deve ter sido substituído
        jan_tons = result[
            (result["Date"].dt.month == 1) & (result["Date"].dt.year == 2026)
        ]["Tons"].tolist()
        self.assertNotIn(45000, jan_tons)
        self.assertIn(46000, jan_tons)

    def test_mantém_dados_fora_do_periodo(self):
        db = _sample_db()
        novo = _sample_novo()
        result = merge_com_banco(novo, db)

        # Dezembro 2025 deve estar intacto
        dez = result[result["Date"].dt.month == 12]
        self.assertEqual(len(dez), 2)

    def test_resultado_ordenado_por_data(self):
        db = _sample_db()
        novo = _sample_novo()
        result = merge_com_banco(novo, db)
        dates = result["Date"].tolist()
        self.assertEqual(dates, sorted(dates))

    def test_month_year_consistentes(self):
        db = _sample_db()
        novo = _sample_novo()
        result = merge_com_banco(novo, db)
        for _, row in result.iterrows():
            self.assertEqual(row["Month"], row["Date"].month)
            self.assertEqual(row["Year"], row["Date"].year)

    def test_banco_vazio(self):
        db = pd.DataFrame(columns=["Date", "Destination", "Origin", "Cargo",
                                   "Tons", "Month", "Year"])
        db["Date"] = pd.to_datetime(db["Date"])
        novo = _sample_novo()
        result = merge_com_banco(novo, db)
        self.assertEqual(len(result), len(novo))

    def test_arquivo_novo_vazio_nao_remove_dados(self):
        db = _sample_db()
        novo = pd.DataFrame(columns=["Date", "Destination", "Origin", "Cargo",
                                     "Tons", "Month", "Year"])
        novo["Date"] = pd.to_datetime(novo["Date"])
        result = merge_com_banco(novo, db)
        self.assertEqual(len(result), len(db))


class TestSalvarLocal(unittest.TestCase):

    def test_salva_arquivo_e_cria_sheet_data_base(self):
        """
        Deixa salvar_local escrever um .xlsx real num diretório temporário
        e verifica o resultado observável: arquivo existe + sheet 'data_base' presente.
        Não mocka ExcelWriter — o pandas ignora esse mock internamente.
        """
        import tempfile
        import openpyxl

        df = _sample_db()

        with tempfile.TemporaryDirectory() as tmpdir:
            target_path = Path(tmpdir) / "test.xlsx"
            salvar_local(df, target_path)

            # Arquivo deve ter sido criado
            self.assertTrue(target_path.exists())

            # Sheet 'data_base' deve existir com os dados corretos
            wb = openpyxl.load_workbook(target_path)
            self.assertIn("data_base", wb.sheetnames)
            ws = wb["data_base"]
            # Cabeçalho deve ter todas as colunas esperadas
            headers = [cell.value for cell in ws[1]]
            for col in ["Date", "Destination", "Origin", "Cargo", "Tons", "Month", "Year"]:
                self.assertIn(col, headers)
            # Deve ter as linhas de dados (1 cabeçalho + 3 linhas)
            self.assertEqual(ws.max_row, 4)


class TestSalvarOnedrive(unittest.TestCase):

    def test_cria_cinco_sheets(self):
        df = _sample_db()
        df["Date"] = pd.to_datetime(df["Date"])
        sheets_criadas = []

        mock_writer = MagicMock()
        mock_writer.__enter__ = Mock(return_value=mock_writer)
        mock_writer.__exit__ = Mock(return_value=False)

        original_to_excel = pd.DataFrame.to_excel

        def fake_to_excel(self_df, writer, sheet_name=None, **kwargs):
            sheets_criadas.append(sheet_name)

        with patch("database.pd.ExcelWriter", return_value=mock_writer), \
             patch("database.Path.mkdir"), \
             patch.object(pd.DataFrame, "to_excel", fake_to_excel):
            salvar_onedrive(df, Path("onedrive/test.xlsx"))

        self.assertIn("data_base", sheets_criadas)
        self.assertIn("2025", sheets_criadas)
        self.assertIn("2026", sheets_criadas)
        self.assertIn("Pivot_2025", sheets_criadas)
        self.assertIn("Pivot_2026", sheets_criadas)


class TestSalvarSqlServer(unittest.TestCase):

    def test_delete_e_insert_sao_chamados(self):
        df = _sample_db()
        df["Date"] = pd.to_datetime(df["Date"])

        mock_cursor = MagicMock()
        mock_cursor.fast_executemany = False
        mock_conn = MagicMock()
        mock_conn.cursor.return_value = mock_cursor

        # Patcha o pyodbc dentro do namespace de 'database', não o módulo global
        mock_pyodbc = MagicMock()
        mock_pyodbc.connect.return_value = mock_conn

        with patch.dict("sys.modules", {"pyodbc": mock_pyodbc}), \
             patch("database.pyodbc", mock_pyodbc):
            salvar_sql_server(df, "SERVER", "DATABASE", "TABLE")

        delete_calls = [
            c for c in mock_cursor.execute.call_args_list
            if "DELETE" in str(c)
        ]
        self.assertTrue(len(delete_calls) >= 1)

        mock_cursor.executemany.assert_called_once()
        insert_sql = mock_cursor.executemany.call_args[0][0]
        self.assertIn("INSERT INTO", insert_sql)
        mock_conn.commit.assert_called_once()

    def test_rollback_em_caso_de_erro(self):
        df = _sample_db()
        df["Date"] = pd.to_datetime(df["Date"])

        mock_cursor = MagicMock()
        mock_cursor.fast_executemany = False
        mock_cursor.executemany.side_effect = Exception("DB error")
        mock_conn = MagicMock()
        mock_conn.cursor.return_value = mock_cursor

        mock_pyodbc = MagicMock()
        mock_pyodbc.connect.return_value = mock_conn

        with patch.dict("sys.modules", {"pyodbc": mock_pyodbc}), \
             patch("database.pyodbc", mock_pyodbc):
            with self.assertRaises(Exception):
                salvar_sql_server(df, "SERVER", "DATABASE", "TABLE")

        mock_conn.rollback.assert_called_once()


# ===========================================================================
# TESTES — downloader.py
# ===========================================================================

class TestExtractServerFilename(unittest.TestCase):

    def _mock_response(self, cd_header: str) -> Mock:
        r = Mock()
        r.headers = {"Content-Disposition": cd_header}
        return r

    def test_filename_simples(self):
        r = self._mock_response('attachment; filename="Sailed Vessels_2026-01-01.xlsx"')
        self.assertEqual(_extract_server_filename(r), "Sailed Vessels_2026-01-01.xlsx")

    def test_filename_sem_aspas(self):
        r = self._mock_response("attachment; filename=Sailed Vessels_2026-01-01.xlsx")
        self.assertEqual(_extract_server_filename(r), "Sailed Vessels_2026-01-01.xlsx")

    def test_filename_rfc5987(self):
        r = self._mock_response(
            "attachment; filename*=UTF-8''Sailed%20Vessels_2026-01-01.xlsx"
        )
        self.assertEqual(_extract_server_filename(r), "Sailed Vessels_2026-01-01.xlsx")

    def test_sem_header_retorna_none(self):
        r = Mock()
        r.headers = {}
        self.assertIsNone(_extract_server_filename(r))

    def test_header_vazio_retorna_none(self):
        r = self._mock_response("")
        self.assertIsNone(_extract_server_filename(r))


class TestBuildOutputName(unittest.TestCase):

    def test_com_data_no_nome_do_servidor(self):
        result = _build_output_name(
            "vessels_sailed_update", "Sailed Vessels_2026-01-01.xlsx"
        )
        self.assertEqual(result, "vessels_sailed_update_Sailed Vessels_2026-01-01.xlsx")

    def test_sem_nome_do_servidor_retorna_padrao(self):
        result = _build_output_name("vessels_sailed_update", None)
        self.assertEqual(result, "vessels_sailed_update.xlsx")

    def test_nome_servidor_sem_data_retorna_padrao(self):
        result = _build_output_name("vessels_sailed_update", "Sailed Vessels.xlsx")
        self.assertEqual(result, "vessels_sailed_update.xlsx")


class TestDownloadFile(unittest.TestCase):

    def _make_fake_downloaded_file(self, tmpdir: str, name: str) -> Path:
        """Cria um .xlsx falso com magic bytes PK válidos na pasta temporária."""
        p = Path(tmpdir) / name
        p.write_bytes(b"PK\x03\x04" + b"\x00" * 5000)
        return p

    def _make_mock_driver(self, downloaded_file: Path):
        """Retorna um mock de webdriver que não faz nada (get() é no-op)."""
        driver = MagicMock()
        driver.get = Mock()
        driver.quit = Mock()
        return driver

    def test_salva_arquivo_com_nome_enriquecido(self):
        import tempfile

        with tempfile.TemporaryDirectory() as dl_tmp, \
             tempfile.TemporaryDirectory() as dest_tmp:

            # Simula o arquivo que o Chrome teria baixado
            downloaded = self._make_fake_downloaded_file(
                dl_tmp, "Sailed Vessels_2026-03-01.xlsx"
            )

            mock_driver = self._make_mock_driver(downloaded)

            # Faz tempfile.mkdtemp retornar nossa pasta controlada
            with patch("downloader.tempfile.mkdtemp", return_value=dl_tmp), \
                 patch("selenium.webdriver.Chrome", return_value=mock_driver), \
                 patch("selenium.webdriver.chrome.service.Service"), \
                 patch("webdriver_manager.chrome.ChromeDriverManager"):
                result = download_file(
                    url="http://fake.url/file",
                    file_name="vessels_sailed_update.xlsx",
                    destination_path=Path(dest_tmp),
                    timeout=5,
                )

        self.assertIn("Sailed Vessels_2026-03-01", str(result))
        self.assertTrue(result.exists())

    def test_levanta_erro_quando_download_nao_aparece(self):
        """Sem arquivos na pasta temp → TimeoutError após timeout curto."""
        import tempfile

        with tempfile.TemporaryDirectory() as dl_tmp, \
             tempfile.TemporaryDirectory() as dest_tmp:

            mock_driver = MagicMock()
            mock_driver.get = Mock()
            mock_driver.quit = Mock()

            # Pasta vazia — _wait_for_download vai esgotar o timeout
            with patch("downloader.tempfile.mkdtemp", return_value=dl_tmp), \
                 patch("selenium.webdriver.Chrome", return_value=mock_driver), \
                 patch("selenium.webdriver.chrome.service.Service"), \
                 patch("webdriver_manager.chrome.ChromeDriverManager"), \
                 patch("downloader.time.sleep"):  # acelera o polling
                with self.assertRaises(TimeoutError):
                    download_file(
                        url="http://fake.url/file",
                        file_name="file.xlsx",
                        destination_path=Path(dest_tmp),
                        timeout=0,  # timeout imediato
                    )


class TestValidateExcelFile(unittest.TestCase):

    def test_arquivo_valido_nao_levanta(self):
        import tempfile
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            # Escreve assinatura ZIP válida + padding para passar o tamanho mínimo
            f.write(b"PK" + b"\x00" * 5000)
            tmp = Path(f.name)
        try:
            from downloader import _validate_excel_file
            _validate_excel_file(tmp)  # não deve levantar
        finally:
            tmp.unlink(missing_ok=True)

    def test_arquivo_html_levanta_valor_error(self):
        import tempfile
        from downloader import _validate_excel_file
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            # Simula página HTML retornada em vez do Excel
            f.write(b"<html><body>Redirecting...</body></html>" * 200)
            tmp = Path(f.name)
        try:
            with self.assertRaises(ValueError) as ctx:
                _validate_excel_file(tmp)
            self.assertIn("ZIP", str(ctx.exception))
        finally:
            tmp.unlink(missing_ok=True)

    def test_arquivo_muito_pequeno_levanta_valor_error(self):
        import tempfile
        from downloader import _validate_excel_file
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            f.write(b"PK" + b"\x00" * 10)  # assinatura OK mas tamanho insuficiente
            tmp = Path(f.name)
        try:
            with self.assertRaises(ValueError) as ctx:
                _validate_excel_file(tmp)
            self.assertIn("KB", str(ctx.exception))
        finally:
            tmp.unlink(missing_ok=True)


# ===========================================================================
# TESTES — latest_file.py
# ===========================================================================

class TestGetLatestFile(unittest.TestCase):

    def test_retorna_arquivo_mais_recente(self):
        """
        Usa um diretório temporário real com arquivos reais.
        Mocka só os ctimes para controlar a ordem sem depender do filesystem.
        """
        import tempfile

        with tempfile.TemporaryDirectory() as tmpdir:
            path_a = Path(tmpdir) / "a.xlsx"
            path_b = Path(tmpdir) / "b.xlsx"
            path_c = Path(tmpdir) / "c.xlsx"

            for p in (path_a, path_b, path_c):
                p.write_bytes(b"data")

            ctime_map = {
                str(path_a): 1000.0,
                str(path_b): 2000.0,
                str(path_c): 3000.0,
            }

            def fake_ctime(p):
                return ctime_map[str(p)]

            with patch("latest_file.os.path.getctime", side_effect=fake_ctime):
                result = get_latest_file(Path(tmpdir))

        self.assertEqual(result.name, "c.xlsx")

    def test_diretorio_inexistente_levanta_erro(self):
        with self.assertRaises(FileNotFoundError):
            get_latest_file(Path("/caminho/que/nao/existe_xyz_abc"))

    def test_diretorio_vazio_levanta_erro(self):
        import tempfile

        with tempfile.TemporaryDirectory() as tmpdir:
            with self.assertRaises(FileNotFoundError):
                get_latest_file(Path(tmpdir))


# ===========================================================================
# Runner
# ===========================================================================

if __name__ == "__main__":
    unittest.main(verbosity=2)